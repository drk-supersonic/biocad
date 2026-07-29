from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from .models import Task

COLUMNS = ["задача", "описание", "исполнитель", "длительность", "предшественники"]


def parse_excel(file_bytes: bytes) -> list[dict]:
    """Парсит Excel в формате из ТЗ и возвращает список "сырых" задач
    (ещё без id/расписания) — по имени, с предшественниками как списком имён.

    Читает напрямую через openpyxl, без pandas: для одного простого листа
    с 5 колонками pandas — лишняя зависимость (тянет за собой numpy, заметно
    увеличивает время установки и холодный старт на бесплатном Render-плане).
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)

    try:
        header = next(rows_iter)
    except StopIteration:
        raise ValueError("Файл пустой")

    col_index = {
        str(c).strip().lower(): i for i, c in enumerate(header) if c is not None
    }
    missing = [c for c in COLUMNS if c not in col_index]
    if missing:
        raise ValueError(f"В файле отсутствуют колонки: {', '.join(missing)}")

    def cell(row: tuple, name: str):
        i = col_index[name]
        return row[i] if i < len(row) else None

    tasks: list[dict] = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue  # пропускаем полностью пустые строки в конце листа

        name = cell(row, "задача")
        if name is None or str(name).strip() == "":
            continue

        raw_pred = cell(row, "предшественники")
        if raw_pred is None or str(raw_pred).strip() == "":
            predecessors: list[str] = []
        else:
            predecessors = [p.strip() for p in str(raw_pred).split(",") if p.strip()]

        raw_desc = cell(row, "описание")
        raw_assignee = cell(row, "исполнитель")
        raw_duration = cell(row, "длительность")

        tasks.append({
            "name": str(name).strip(),
            "description": "" if raw_desc is None else str(raw_desc).strip(),
            "assignee": "" if raw_assignee is None else str(raw_assignee).strip(),
            "duration": int(raw_duration) if raw_duration is not None else 1,
            "predecessors": predecessors,
        })
    return tasks


def build_excel(tasks: list[Task]) -> bytes:
    """Собирает Excel обратно в исходном формате колонок из ТЗ.
    Зависимости экспортируются по именам задач (человекочитаемо)."""
    by_id = {t.id: t for t in tasks}

    wb = Workbook()
    ws = wb.active
    ws.title = "План"
    ws.append(COLUMNS + ["начало", "окончание"])

    for t in tasks:
        pred_names = [by_id[p].name for p in t.predecessors if p in by_id]
        ws.append([
            t.name,
            t.description,
            t.assignee,
            t.duration,
            ", ".join(pred_names),
            t.start.isoformat() if t.start else "",
            t.finish.isoformat() if t.finish else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
